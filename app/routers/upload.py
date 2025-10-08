from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook

from app.auth import get_current_user
from app.config import settings
from app.schemas import UploadResponse, User
from app.services.storage import data_store

router = APIRouter()


def _normalize_header(header: str) -> str:
    return header.strip().lower().replace(" ", "_")


def _infer_columns(headers: List[str]) -> Dict[str, Optional[str]]:
    normalized = {h: _normalize_header(h) for h in headers}
    section_col = None
    tech_col = None
    for original, norm in normalized.items():
        if section_col is None and any(h in norm for h in settings.section_column_hints):
            section_col = original
        if tech_col is None and any(h in norm for h in settings.technology_column_hints):
            tech_col = original
    return {"section": section_col, "technology": tech_col}


@router.post("/excel", response_model=UploadResponse, summary="Upload Excel knowledge base")
async def upload_excel(
    file: UploadFile = File(..., description="Excel .xlsx file"),
    current_user: User = Depends(get_current_user),
) -> UploadResponse:
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    try:
        wb = load_workbook(filename=file.file, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read workbook: {exc}")

    total_rows = 0
    inferred_columns_agg: Dict[str, Optional[str]] = {"section": None, "technology": None}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers_row = rows[0]
        headers = [str(h) if h is not None else "" for h in headers_row]
        inferred = _infer_columns(headers)
        # aggregate first found
        if inferred_columns_agg["section"] is None and inferred["section"] is not None:
            inferred_columns_agg["section"] = inferred["section"]
        if inferred_columns_agg["technology"] is None and inferred["technology"] is not None:
            inferred_columns_agg["technology"] = inferred["technology"]

        for idx, row in enumerate(rows[1:], start=2):
            record = {headers[i]: row[i] for i in range(len(headers))}
            data_store.add_row(
                record=record,
                sheet_name=ws.title,
                section_column=inferred["section"],
                technology_column=inferred["technology"],
                source_filename=file.filename,
                row_index=idx,
            )
            total_rows += 1

    return UploadResponse(
        filename=file.filename,
        worksheets_loaded=len(wb.worksheets),
        rows_indexed=total_rows,
        inferred_columns=inferred_columns_agg,
    )


@router.get("/status", summary="Get data store status")
async def status(current_user: User = Depends(get_current_user)) -> Dict[str, int]:
    return {"rows_indexed": data_store.count_rows()}
