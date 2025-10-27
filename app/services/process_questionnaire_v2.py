from app.services.excel_processor import excel_to_table_text
from app.services.excel_parser_v3 import ExcelParserV3
from pathlib import Path
import json
import os
from dotenv import load_dotenv
from app.services.vector_db import search_qdrant
from openpyxl.utils import get_column_letter
from app.config import settings
from app.services.langchain_bedrock import bedrock_client

import openpyxl

load_dotenv()


def answer_question_with_rag_v2(
    question_data: dict, table_text: str, context: str
) -> str:
    """
    Answer a single question using RAG + LLM with detailed response instructions.
    
    Args:
        question_data: Dictionary containing question_text, cell_location, and response_instruction
        table_text: The full table text from Excel for context
        context: Relevant knowledge context from vector search
    
    Returns:
        Cell reference and value in format "CellRef = Value"
    """
    
    question_text = question_data["question_text"]
    cell_location = question_data["cell_location"]
    response_instruction = question_data["response_instruction"]
    
    print(f"Processing question: {question_text[:100]}...")
    print(f"Cell location: {cell_location}")
    
    system_message = """You are a Compass Group compliance team member filling out a client questionnaire. Answer naturally and concisely.

CRITICAL INSTRUCTIONS:
1. Keep answers SHORT and CRISP - maximum 2-3 sentences
2. Answer ONLY with information explicitly provided in the context
3. Sound like a real Compass Group employee, not AI
4. If context is insufficient, leave blank - do not guess or assume
5. Use simple, direct language
6. Follow response instructions exactly

For the question, follow this process:
1. <THINKING>What specific information is needed for this question?</THINKING>
2. <CONTEXT_SEARCH>What relevant information exists in the provided context?</CONTEXT_SEARCH>
3. <VALIDATION>Is the context information sufficient and directly relevant?</VALIDATION>
4. <DECISION>Should I answer or leave blank?</DECISION>

Output format:
- If answering: <CellRef> = "<Value>"
- If no answer: <CellRef> = ""

RESPONSE GUIDELINES:
- Keep responses SHORT and DIRECT
- Use only information explicitly stated in the context
- Sound natural and professional, like a Compass Group employee
- Avoid phrases like "based on the context", "as mentioned", "according to"
- Do not use AI knowledge about Compass Group
- If context doesn't contain the answer, leave blank
- Do not infer, assume, or add information not in context
- Use simple language without technical jargon unless it's in the context
- Return only the cell reference and value, nothing else"""

    prompt = f"""Question: {question_text}
Cell location: {cell_location}
Response instruction: {response_instruction}

Table context:
{table_text}

Knowledge context:
{context}

Answer this question using ONLY the information provided above. Keep your answer short and direct. If the information is not available in the context, leave the cell blank."""

    # Prepare metadata for tracing
    metadata = {
        "question_text": question_text[:100],
        "cell_location": cell_location,
        "context_length": len(context),
        "table_text_length": len(table_text)
    }

    # Call Claude via LangChain with tracing
    claude_response = bedrock_client.invoke_with_tracing(
        prompt=prompt,
        system_message=system_message
    )
    
    # Get cost estimate
    cost_estimate = bedrock_client.get_cost_estimate(prompt, claude_response)
    print(f"Cost estimate for question: ${cost_estimate['total_cost_usd']:.6f}")
    print(f"Tokens - Input: {cost_estimate['input_tokens']:.0f}, Output: {cost_estimate['output_tokens']:.0f}")
    
    print(f"LLM response (with thinking): {claude_response}")
    
    # Clean the response by removing thinking tags
    cleaned_response = clean_thinking_tags(claude_response)
    print(f"Cleaned response: {cleaned_response}")
    
    return cleaned_response


def clean_thinking_tags(response: str) -> str:
    """
    Remove thinking tags from the response while preserving the table format.
    """
    import re
    
    # Remove all thinking tags and their content
    # Pattern matches: <TAG_NAME>content</TAG_NAME>
    thinking_pattern = r'<THINKING>.*?</THINKING>'
    context_search_pattern = r'<CONTEXT_SEARCH>.*?</CONTEXT_SEARCH>'
    validation_pattern = r'<VALIDATION>.*?</VALIDATION>'
    decision_pattern = r'<DECISION>.*?</DECISION>'
    reasoning = r'<reasoning>.*?</reasoning>'
    
    # Remove all thinking tags
    cleaned = re.sub(thinking_pattern, '', response, flags=re.DOTALL)
    cleaned = re.sub(context_search_pattern, '', cleaned, flags=re.DOTALL)
    cleaned = re.sub(validation_pattern, '', cleaned, flags=re.DOTALL)
    cleaned = re.sub(decision_pattern, '', cleaned, flags=re.DOTALL)
    cleaned = re.sub(reasoning, '', cleaned, flags=re.DOTALL)
    
    # Clean up extra whitespace and empty lines
    lines = cleaned.split('\n')
    cleaned_lines = []
    
    for line in lines:
        line = line.strip()
        if line and not line.startswith('<') and not line.endswith('>'):
            cleaned_lines.append(line)
    
    return '\n'.join(cleaned_lines)


def write_back_to_excel_v2(file_path: str, output_path: str, llm_filled_text: str):
    """Update the original Excel file with filled answers from LLM."""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    merged_map = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row = merged_range.min_col, merged_range.min_row
        top_left = f"{get_column_letter(min_col)}{min_row}"
        for row in range(min_row, merged_range.max_row + 1):
            for col in range(min_col, merged_range.max_col + 1):
                merged_map[f"{get_column_letter(col)}{row}"] = top_left

    for line in llm_filled_text.splitlines():
        if "=" not in line:
            continue
        cell_ref, value = line.split("=", 1)
        cell_ref = cell_ref.strip()
        value = value.strip().strip('"')
        target_cell = merged_map.get(cell_ref, cell_ref)
        if sheet[target_cell].value in [None, ""]:  # only overwrite empty
            sheet[target_cell] = value

    wb.save(output_path)


def process_questionnaire_v2(file_path: str, save_thinking_process: bool = False):
    """
    Process questionnaire using excel_parser_v3 structure with individual question processing.
    
    Args:
        file_path: Path to the Excel questionnaire file
        save_thinking_process: Whether to save the thinking process for debugging (default: False)
    """
    print(f"🚀 Starting questionnaire processing V2 for: {file_path}")
    print(f"🧠 Using excel_parser_v3 structure with individual question processing")
    
    # Step 1: Convert Excel to table text
    table_text = excel_to_table_text(file_path, historical=False)
    print("Excel converted to table text.")
    
    # Step 2: Use excel_parser_v3 to get structure
    parser_v3 = ExcelParserV3()
    structure = parser_v3.analyze_table_and_generate_instructions(table_text)
    print(f"Structure identified using excel_parser_v3: {len(structure.get('questions', []))} questions found")
    
    if "error" in structure:
        print(f"❌ Error in structure analysis: {structure['error']}")
        return None
    
    questions = structure.get("questions", [])
    if not questions:
        print("❌ No questions found in the structure")
        return None
    
    all_filled = []
    all_thinking_processes = []  # Store thinking processes for debugging
    total_cost = 0.0
    total_input_tokens = 0
    total_output_tokens = 0
    
    print(f"📊 Processing {len(questions)} questions individually...")
    
    for i, question_data in enumerate(questions, 1):
        print(f"🔄 Processing question {i}/{len(questions)}")
        
        # Get context for this specific question
        context_results = search_qdrant(question_data["question_text"])
        
        # Extract text and scores from context results
        context_texts = [result["text"] for result in context_results]
        context_scores = [result["score"] for result in context_results]
        context = "\n\n".join(context_texts)
        
        print(f"Context scores for question {i}: {context_scores}")
        
        # Get the raw response with thinking tags
        raw_response = answer_question_with_rag_v2(question_data, table_text, context)
        
        # Clean the response for Excel processing
        filled = clean_thinking_tags(raw_response)
        all_filled.append(filled)
        
        # Store thinking process if requested
        if save_thinking_process:
            all_thinking_processes.append({
                "question": i,
                "question_text": question_data["question_text"],
                "cell_location": question_data["cell_location"],
                "thinking_process": raw_response
            })
        
        # Estimate cost for this question (simplified)
        batch_cost = bedrock_client.get_cost_estimate("", filled)
        total_cost += batch_cost['total_cost_usd']
        total_input_tokens += batch_cost['input_tokens']
        total_output_tokens += batch_cost['output_tokens']
    
    combined_filled = "\n".join(all_filled)
    output_path = Path(file_path).with_name("filled_v2_1_" + Path(file_path).name)
    write_back_to_excel_v2(file_path, output_path, combined_filled)

    # Save thinking process if requested
    if save_thinking_process and all_thinking_processes:
        thinking_path = Path(file_path).with_name("thinking_process_v2_" + Path(file_path).stem + ".txt")
        with open(thinking_path, "w", encoding="utf-8") as f:
            for process in all_thinking_processes:
                f.write(f"=== QUESTION {process['question']} ===\n")
                f.write(f"Question: {process['question_text']}\n")
                f.write(f"Cell Location: {process['cell_location']}\n")
                f.write(f"Thinking Process:\n{process['thinking_process']}\n")
                f.write("\n" + "="*80 + "\n\n")
        print(f"🧠 Thinking process saved to: {thinking_path}")

    # Print summary
    print(f"\n📈 Processing Summary:")
    print(f"   Total questions processed: {len(questions)}")
    print(f"   Total estimated cost: ${total_cost:.6f}")
    print(f"   Total input tokens: {total_input_tokens:.0f}")
    print(f"   Total output tokens: {total_output_tokens:.0f}")
    print(f"   🧠 Individual question processing with detailed response instructions")
    print(f"   Output saved to: {output_path}")

    return output_path


if __name__ == "__main__":
    file_path = "complex_format_empty.xlsx"
    # Process with excel_parser_v3 structure and save thinking process for debugging
    output_path = process_questionnaire_v2(file_path, save_thinking_process=True)
    if output_path:
        print(f"The questionnaire is filled and is in the {output_path}")
    else:
        print("Processing failed.")
