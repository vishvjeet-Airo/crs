from app.services.excel_processor import excel_to_table_text
from app.services.excel_parser import identify_excel_structure
from pathlib import Path
import json
import os
from dotenv import load_dotenv
from app.services.get_rows import get_rows
from app.services.extract_questions import extract_questions
from app.services.vector_db import search_qdrant
from openpyxl.utils import get_column_letter
from app.config import settings
from app.services.langchain_bedrock import bedrock_client

import openpyxl

load_dotenv()


def answer_batch_with_rag(
    batch_rows: list, structure: dict, extracted_rows: str
) -> str:
    """
    Given batch rows, query RAG + LLM to fill answers.
    Returns filled table-text format with LangSmith tracing and cost monitoring.
    """

    row_wise_text = get_rows(extracted_rows, batch_rows)
    questions = extract_questions(row_wise_text, structure["columns"]["Question"])
    batch_query = "\n".join(questions)
    context_results = search_qdrant(batch_query)
    
    # Extract text and scores from context results
    context_texts = [result["text"] for result in context_results]
    context_scores = [result["score"] for result in context_results]
    context = "\n\n".join(context_texts)
    
    print(f"context for batch with row nums - {batch_rows} is : {context}")
    print(f"context scores: {context_scores}")

    system_message = """You are a compliance team member at Compass Group filling out a client questionnaire. You must think through EACH INDIVIDUAL QUESTION step by step to avoid hallucination.

CRITICAL INSTRUCTIONS:
1. Process EACH question in the batch individually with full thinking process
2. Check headers for response instructions, options, and format requirements
3. Answer like a human compliance professional, not AI-generated
4. Use natural language that sounds like a real compliance team response
5. If options are provided, choose from those options only
6. Add meaningful comments that a compliance team would write

For EACH question, follow this exact process:
1. <THINKING>Analyze this specific question and identify what information is needed</THINKING>
2. <HEADER_CHECK>Check if headers contain response instructions, options, or format requirements</HEADER_CHECK>
3. <CONTEXT_SEARCH>Search through the provided context for information that directly answers this question</CONTEXT_SEARCH>
4. <VALIDATION>Validate that the information found is sufficient and directly relevant to this question</VALIDATION>
5. <DECISION>Decide whether to provide an answer or leave blank based on context sufficiency</DECISION>

Output format for EACH question:
- For each cell that needs filling, use this exact format:
  <THINKING>Your reasoning for this specific question</THINKING>
  <HEADER_CHECK>Any header instructions or options found</HEADER_CHECK>
  <CONTEXT_SEARCH>Relevant context snippets for this question</CONTEXT_SEARCH>
  <VALIDATION>Your validation of the context for this question</VALIDATION>
  <DECISION>Your decision and reasoning for this question</DECISION>
  <CellRef> = "<Value>"

- If no answer can be provided based on context:
  <THINKING>Question analysis here</THINKING>
  <HEADER_CHECK>Header check results here</HEADER_CHECK>
  <CONTEXT_SEARCH>Context search results here</CONTEXT_SEARCH>
  <VALIDATION>Why context is insufficient here</VALIDATION>
  <DECISION>Decision to leave blank here</DECISION>
  <CellRef> = ""

RESPONSE GUIDELINES:
- Use the following format for **every filled cell**:
  <CellRef> = "<Value>"
  Example: D30 = "Yes" or D30 = "AES-256 encryption"
- If a cell has no answer, output it as:
  <CellRef> = ""
- Answer like a human compliance professional would
- Use natural, professional language
- If options are provided in headers, choose from those options
- Add meaningful comments that provide value
- Do not use AI-like phrases like "based on the context" or "as mentioned"
- Be specific and actionable in responses
- Do not output tabs, extra spaces, or any markdown formatting
- Keep original cell references (like C30, D30, etc)
- Only fill where required, do not change existing filled data
- Only provide a response in a cell if it is explicitly supported by the given context
- If the context does not contain enough information to answer a question, leave the corresponding cell(s) completely blank
- Do not assume, infer, or guess
- Do not introduce new information that is not present in the context
- Return only the filled-table text in the format above and nothing else, as this will be parsed to fill it in Excel."""

    prompt = f"""You are provided:
1. The Excel questionnaire structure: {json.dumps(structure)}
2. The following rows need to be filled: {row_wise_text}
3. Relevant knowledge context:
{context}

IMPORTANT: Process each question individually. For each question in the batch, go through the complete thinking process (THINKING, HEADER_CHECK, CONTEXT_SEARCH, VALIDATION, DECISION) before providing your answer."""

    # Prepare metadata for tracing
    metadata = {
        "batch_rows": batch_rows,
        "question_count": len(questions),
        "context_length": len(context),
        "structure_columns": list(structure.get("columns", {}).keys())
    }

    # Call Claude via LangChain with tracing
    claude_response = bedrock_client.invoke_with_tracing(
        prompt=prompt,
        system_message=system_message
    )
    
    # Get cost estimate
    cost_estimate = bedrock_client.get_cost_estimate(prompt, claude_response)
    print(f"Cost estimate for batch {batch_rows}: ${cost_estimate['total_cost_usd']:.6f}")
    print(f"Tokens - Input: {cost_estimate['input_tokens']:.0f}, Output: {cost_estimate['output_tokens']:.0f}")
    
    print(f"llm response (with thinking): {claude_response}")
    
    # Clean the response by removing thinking tags
    cleaned_response = clean_thinking_tags(claude_response)
    print(f"cleaned response: {cleaned_response}")
    
    return cleaned_response


def clean_thinking_tags(response: str) -> str:
    """
    Remove thinking tags from the response while preserving the table format.
    """
    import re
    
    # Remove all thinking tags and their content
    # Pattern matches: <TAG_NAME>content</TAG_NAME>
    thinking_pattern = r'<THINKING>.*?</THINKING>'
    header_check_pattern = r'<HEADER_CHECK>.*?</HEADER_CHECK>'
    context_search_pattern = r'<CONTEXT_SEARCH>.*?</CONTEXT_SEARCH>'
    validation_pattern = r'<VALIDATION>.*?</VALIDATION>'
    decision_pattern = r'<DECISION>.*?</DECISION>'
    reasoning = r'<reasoning>.*?</reasoning>'
    
    # Remove all thinking tags
    cleaned = re.sub(thinking_pattern, '', response, flags=re.DOTALL)
    cleaned = re.sub(header_check_pattern, '', cleaned, flags=re.DOTALL)
    cleaned = re.sub(context_search_pattern, '', cleaned, flags=re.DOTALL)
    cleaned = re.sub(validation_pattern, '', cleaned, flags=re.DOTALL)
    cleaned = re.sub(decision_pattern, '', cleaned, flags=re.DOTALL)
    cleaned = re.sub(reasoning, '', cleaned, flags=re.DOTALL)
    
    # Clean up extra whitespace and empty lines
    lines = cleaned.split('\n')
    cleaned_lines = []
    
    for line in lines:
        line = line.strip()
        if line and not line.startswith('<') and not line.endswith('>'):
            cleaned_lines.append(line)
    
    return '\n'.join(cleaned_lines)


def write_back_to_excel(file_path: str, output_path: str, llm_filled_text: str):
    """Update the original Excel file with filled answers from LLM."""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    merged_map = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row = merged_range.min_col, merged_range.min_row
        top_left = f"{get_column_letter(min_col)}{min_row}"
        for row in range(min_row, merged_range.max_row + 1):
            for col in range(min_col, merged_range.max_col + 1):
                merged_map[f"{get_column_letter(col)}{row}"] = top_left

    for line in llm_filled_text.splitlines():
        if "=" not in line:
            continue
        cell_ref, value = line.split("=", 1)
        cell_ref = cell_ref.strip()
        value = value.strip().strip('"')
        target_cell = merged_map.get(cell_ref, cell_ref)
        if sheet[target_cell].value in [None, ""]:  # only overwrite empty
            sheet[target_cell] = value

    wb.save(output_path)


def process_questionnaire(file_path: str, save_thinking_process: bool = False):
    """
    Process questionnaire with LangSmith tracing, cost monitoring, and thinking-out-loud approach.
    
    Args:
        file_path: Path to the Excel questionnaire file
        save_thinking_process: Whether to save the thinking process for debugging (default: False)
    """
    print(f"🚀 Starting questionnaire processing for: {file_path}")
    print(f"🧠 Thinking-out-loud mode: Enabled (reduces hallucination)")
    
    extracted = excel_to_table_text(file_path, historical=False)
    print("excel is converted to a text.", extracted)
    structure = identify_excel_structure(extracted)
    print(f"structure is identified: {structure}")

    all_filled = []
    all_thinking_processes = []  # Store thinking processes for debugging
    total_cost = 0.0
    total_input_tokens = 0
    total_output_tokens = 0
    
    print(f"📊 Processing {len(structure['batches'])} batches...")
    
    for i, batch in enumerate(structure["batches"], 1):
        print(f"🔄 Processing batch {i}/{len(structure['batches'])} (rows: {batch['rows']})")
        
        # Get the raw response with thinking tags
        raw_response = answer_batch_with_rag_raw(batch["rows"], structure, extracted)
        
        # Clean the response for Excel processing
        filled = clean_thinking_tags(raw_response)
        all_filled.append(filled)
        
        # Store thinking process if requested
        if save_thinking_process:
            all_thinking_processes.append({
                "batch": i,
                "rows": batch["rows"],
                "thinking_process": raw_response
            })
        
        # Estimate cost for this batch (simplified)
        batch_cost = bedrock_client.get_cost_estimate("", filled)
        total_cost += batch_cost['total_cost_usd']
        total_input_tokens += batch_cost['input_tokens']
        total_output_tokens += batch_cost['output_tokens']
    
    combined_filled = "\n".join(all_filled)
    output_path = Path(file_path).with_name("filled1_" + Path(file_path).name)
    write_back_to_excel(file_path, output_path, combined_filled)

    # Save thinking process if requested
    if save_thinking_process and all_thinking_processes:
        thinking_path = Path(file_path).with_name("thinking_process_" + Path(file_path).stem + ".txt")
        with open(thinking_path, "w", encoding="utf-8") as f:
            for process in all_thinking_processes:
                f.write(f"=== BATCH {process['batch']} (Rows: {process['rows']}) ===\n")
                f.write(process['thinking_process'])
                f.write("\n\n" + "="*80 + "\n\n")
        print(f"🧠 Thinking process saved to: {thinking_path}")

    # Print summary
    print(f"\n📈 Processing Summary:")
    print(f"   Total batches processed: {len(structure['batches'])}")
    print(f"   Total estimated cost: ${total_cost:.6f}")
    print(f"   Total input tokens: {total_input_tokens:.0f}")
    print(f"   Total output tokens: {total_output_tokens:.0f}")
    print(f"   🧠 Thinking-out-loud approach used to reduce hallucination")
    print(f"   Output saved to: {output_path}")

    return output_path


def answer_batch_with_rag_raw(batch_rows: list, structure: dict, extracted_rows: str) -> str:
    """
    Get the raw response with thinking tags (for debugging purposes).
    """
    row_wise_text = get_rows(extracted_rows, batch_rows)
    questions = extract_questions(row_wise_text, structure["columns"]["Question"])
    batch_query = "\n".join(questions)
    context_results = search_qdrant(batch_query)
    
    # Extract text and scores from context results
    context_texts = [result["text"] for result in context_results]
    context_scores = [result["score"] for result in context_results]
    context = "\n\n".join(context_texts)
    
    print(f"context for batch with row nums - {batch_rows} is : {context}")
    print(f"context scores: {context_scores}")

    system_message = """You are a compliance team member at Compass Group filling out a client questionnaire. You must think through EACH INDIVIDUAL QUESTION step by step to avoid hallucination.

CRITICAL INSTRUCTIONS:
1. Process EACH question in the batch individually with full thinking process
2. Check headers for response instructions, options, and format requirements
3. Answer like a human compliance professional, not AI-generated
4. Use natural language that sounds like a real compliance team response
5. If options are provided, choose from those options only
6. Add meaningful comments that a compliance team would write

For EACH question, follow this exact process:
1. <THINKING>Analyze this specific question and identify what information is needed</THINKING>
2. <HEADER_CHECK>Check if headers contain response instructions, options, or format requirements</HEADER_CHECK>
3. <CONTEXT_SEARCH>Search through the provided context for information that directly answers this question</CONTEXT_SEARCH>
4. <VALIDATION>Validate that the information found is sufficient and directly relevant to this question</VALIDATION>
5. <DECISION>Decide whether to provide an answer or leave blank based on context sufficiency</DECISION>

Output format for EACH question:
- For each cell that needs filling, use this exact format:
  <THINKING>Your reasoning for this specific question</THINKING>
  <HEADER_CHECK>Any header instructions or options found</HEADER_CHECK>
  <CONTEXT_SEARCH>Relevant context snippets for this question</CONTEXT_SEARCH>
  <VALIDATION>Your validation of the context for this question</VALIDATION>
  <DECISION>Your decision and reasoning for this question</DECISION>
  <CellRef> = "<Value>"

- If no answer can be provided based on context:
  <THINKING>Question analysis here</THINKING>
  <HEADER_CHECK>Header check results here</HEADER_CHECK>
  <CONTEXT_SEARCH>Context search results here</CONTEXT_SEARCH>
  <VALIDATION>Why context is insufficient here</VALIDATION>
  <DECISION>Decision to leave blank here</DECISION>
  <CellRef> = ""

RESPONSE GUIDELINES:
- Use the following format for **every filled cell**:
  <CellRef> = "<Value>"
  Example: D30 = "Yes" or D30 = "AES-256 encryption"
- If a cell has no answer, output it as:
  <CellRef> = ""
- Answer like a human compliance professional would
- Use natural, professional language
- If options are provided in headers, choose from those options
- Add meaningful comments that provide value
- Do not use AI-like phrases like "based on the context" or "as mentioned"
- Be specific and actionable in responses
- Do not output tabs, extra spaces, or any markdown formatting
- Keep original cell references (like C30, D30, etc)
- Only fill where required, do not change existing filled data
- Only provide a response in a cell if it is explicitly supported by the given context
- If the context does not contain enough information to answer a question, leave the corresponding cell(s) completely blank
- Do not assume, infer, or guess
- Do not introduce new information that is not present in the context
- Return only the filled-table text in the format above and nothing else, as this will be parsed to fill it in Excel."""

    prompt = f"""You are provided:
1. The Excel questionnaire structure: {json.dumps(structure)}
2. The following rows need to be filled: {row_wise_text}
3. Relevant knowledge context:
{context}

IMPORTANT: Process each question individually. For each question in the batch, go through the complete thinking process (THINKING, HEADER_CHECK, CONTEXT_SEARCH, VALIDATION, DECISION) before providing your answer."""

    # Prepare metadata for tracing
    metadata = {
        "batch_rows": batch_rows,
        "question_count": len(questions),
        "context_length": len(context),
        "structure_columns": list(structure.get("columns", {}).keys()),
        "thinking_mode": True
    }

    # Call Claude via LangChain with tracing
    claude_response = bedrock_client.invoke_with_tracing(
        prompt=prompt,
        system_message=system_message
    )
    
    # Get cost estimate
    cost_estimate = bedrock_client.get_cost_estimate(prompt, claude_response)
    print(f"Cost estimate for batch {batch_rows}: ${cost_estimate['total_cost_usd']:.6f}")
    print(f"Tokens - Input: {cost_estimate['input_tokens']:.0f}, Output: {cost_estimate['output_tokens']:.0f}")
    
    return claude_response


if __name__ == "__main__":
    file_path = "complex_format_empty.xlsx"
    # Process with thinking-out-loud approach and save thinking process for debugging
    output_path = process_questionnaire(file_path, save_thinking_process=True)
    print(f"The questionnaire is filled and is in the {output_path}")
